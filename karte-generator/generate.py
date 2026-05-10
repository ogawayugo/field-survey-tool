#!/usr/bin/env python3
"""
街路樹診断カルテ生成スクリプト

PWA「街路樹現場調査」からエクスポートしたJSONファイルを入力として、
渋谷氷川の杜様式の街路樹診断カルテExcelファイルを生成します。

使い方:
    python generate.py survey.json
    python generate.py survey.json --template shibuya
    python generate.py survey.json --output my_karte.xlsx

要件:
    pip install openpyxl pillow

"""

import argparse
import base64
import io
import json
import os
import re
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.worksheet.worksheet import Worksheet
    from PIL import Image
    from photo_annotator import annotate_photo
except ImportError:
    print("ERROR: 必要なライブラリがインストールされていません。", file=sys.stderr)
    print("以下を実行してください: pip install openpyxl pillow", file=sys.stderr)
    sys.exit(1)


# ===================================================================
# 定数
# ===================================================================

# exe化されたかどうかで取得方法を分岐
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = Path(sys.executable).parent
else:
    SCRIPT_DIR = Path(__file__).parent
TEMPLATES_DIR = SCRIPT_DIR / 'templates'

# 半角→全角変換テーブル
HALFWIDTH_TO_FULLWIDTH_DIGITS = str.maketrans('12345', '１２３４５')

# A/B1/B2/C → 全角
ABC_TO_FULLWIDTH = {
    'A': 'Ａ',
    'B1': 'Ｂ１',
    'B2': 'Ｂ２',
    'C': 'Ｃ',
}

# A/B1/B2/C → 活力判定の長文ラベル
ABC_TO_VITALITY_LABEL = {
    'A': '健全か健全に近い',
    'B1': '注意すべき被害が見られる',
    'B2': '著しい被害が見られる',
    'C': '不健全',
}

# 標準診断項目リスト（メモから抽出する時の参照）
KNOWN_DIAGNOSIS_ITEMS = [
    '樹皮枯死・欠損・腐朽',
    '開口空洞(芯に達しない)',
    '開口空洞（芯に達しない）',
    '開口空洞(芯に達する)',
    '開口空洞（芯に達する）',
    'キノコ（子実体）',
    'キノコ',
    '木槌打診異常',
    '分岐部・付根の異常',
    '胴枯れなどの病害',
    '虫穴・虫フン・ヤニ',
    '根元の揺らぎ',
    '鋼棒貫入異常',
    '巻き根',
    'ルートカラー見えない',
    '露出根被害',
    '不自然な傾斜',
    '枯枝',
    'スタブカット',
]


# ===================================================================
# ユーティリティ
# ===================================================================

def log(msg: str):
    """進捗ログ"""
    print(f"  {msg}")


def err(msg: str):
    """エラーログ"""
    print(f"  ERROR: {msg}", file=sys.stderr)


def warn(msg: str):
    """警告ログ"""
    print(f"  WARN: {msg}", file=sys.stderr)


def pt_to_emu(pt: float) -> int:
    """pt → EMU (English Metric Unit)"""
    return int(pt * 12700)


def pt_to_pixel(pt: float, dpi: int = 96) -> int:
    """pt → pixel (96 DPI default)"""
    return int(pt * dpi / 72)


def normalize_text(s: str) -> str:
    """文字列の正規化（全角・半角の括弧を統一）"""
    if not s:
        return ''
    # 全角括弧を半角にして比較しやすくする
    return s.replace('（', '(').replace('）', ')').strip()


def format_date(date_str: str) -> str:
    """ISO形式の日付を「  YYYY年  MM月  DD日」に変換"""
    if not date_str:
        return ''
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', date_str)
    if not m:
        return date_str
    return f"　　{m.group(1)}年　{int(m.group(2))}月　{int(m.group(3))}日"


# ===================================================================
# テンプレート設定の読み込み
# ===================================================================

def load_template_config(template_id: str) -> dict:
    """テンプレート設定JSONを読み込む"""
    config_path = TEMPLATES_DIR / f"{template_id}.json"
    if not config_path.exists():
        raise FileNotFoundError(f"テンプレート設定が見つかりません: {config_path}")

    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_template_excel_path(config: dict) -> Path:
    """テンプレートExcelファイルのパスを返す"""
    return TEMPLATES_DIR / config['template_file']


# ===================================================================
# メモから部位ごとの項目を抽出
# ===================================================================

def parse_memo_to_parts(memo: str) -> dict:
    """
    現場メモを部位ごとに分割する。
    入力例: "根元:キノコ（子実体）、露出根被害5×20cm\n幹:樹皮欠損2×3cm"
    出力: {"根元": ["キノコ（子実体）", "露出根被害5×20cm"], "幹": [...], "大枝": [...]}
    """
    result = {'根元': [], '幹': [], '大枝': [], '_free': []}
    if not memo:
        return result

    lines = [line.strip() for line in memo.split('\n') if line.strip()]
    for line in lines:
        # 「部位:」「部位：」「部位 :」など寛容にマッチ
        m = re.match(r'^(根元|幹|大枝|枝)\s*[:：]\s*(.*)', line)
        if m:
            part = m.group(1)
            if part == '枝':
                part = '大枝'
            content = m.group(2).strip()
            if content:
                # 読点で分割して各項目に
                items = [s.strip() for s in re.split(r'[、,]', content) if s.strip()]
                # 連続した「、、」で出来る空文字列も除去済み
                result[part].extend(items)
        else:
            # 部位ラベルなしの行
            result['_free'].append(line)
    return result


def extract_diagnosis_item_name(text: str) -> str:
    """
    項目名と寸法が混ざった文字列から、項目名部分だけを取り出す
    例: "樹皮枯死・欠損・腐朽5×20cm" → "樹皮枯死・欠損・腐朽"
    例: "開口空洞（芯に達しない）4×5cm" → "開口空洞（芯に達しない）"
    例: "傾斜・小（南方向）" → "不自然な傾斜"（キーワードマッチ）
    """
    if not text:
        return ''
    text = text.strip()
    # 完全一致を優先
    for item in KNOWN_DIAGNOSIS_ITEMS:
        if text == item:
            return item
    # 前方一致
    for item in KNOWN_DIAGNOSIS_ITEMS:
        if text.startswith(item):
            return item
    # キーワードマッチ（項目名が現場での略記とずれる場合の救済）
    keyword_map = {
        '傾斜': '不自然な傾斜',
        '揺らぎ': '根元の揺らぎ',
        '揺れ': '根元の揺らぎ',
        '鋼棒': '鋼棒貫入異常',
        '貫入': '鋼棒貫入異常',
        '巻き根': '巻き根',
        '露出根': '露出根被害',
        '根の切断': '露出根被害',
        '根切断': '露出根被害',
        'ルートカラー': 'ルートカラー見えない',
        '深植': 'ルートカラー見えない',
        '盛土': 'ルートカラー見えない',
        '枯枝': '枯枝',
        'スタブ': 'スタブカット',
        'キノコ': 'キノコ（子実体）',
        '子実体': 'キノコ（子実体）',
        'ベッコウタケ': 'キノコ（子実体）',
        'コフキ': 'キノコ（子実体)',
        '打診': '木槌打診異常',
        '入皮': '分岐部・付根の異常',
        '分岐': '分岐部・付根の異常',
        '胴枯': '胴枯れなどの病害',
        'カミキリ': '虫穴・虫フン・ヤニ',
        '虫穴': '虫穴・虫フン・ヤニ',
        '虫フン': '虫穴・虫フン・ヤニ',
        'ヤニ': '虫穴・虫フン・ヤニ',
    }
    for keyword, mapped in keyword_map.items():
        if keyword in text:
            return mapped
    return ''


# ===================================================================
# シート複製（openpyxlの copy_worksheet を使う）
# ===================================================================

def copy_template_sheet(wb, template_sheet_name: str, new_sheet_name: str) -> Worksheet:
    """
    テンプレートシートを複製する。
    openpyxl の copy_worksheet は同じワークブック内でしか使えないので、
    workbookごとロードしている前提。
    """
    template_sheet = wb[template_sheet_name]
    new_sheet = wb.copy_worksheet(template_sheet)
    new_sheet.title = new_sheet_name
    return new_sheet


# ===================================================================
# データ書き込み
# ===================================================================

def write_basic_info(sheet: Worksheet, tree: dict, survey_meta: dict, config: dict):
    """基本情報の書き込み"""
    bi = config['basic_info']

    mapping = {
        'tree_number': tree.get('treeNumber', ''),
        'species': tree.get('species', ''),
        'height': tree.get('height', ''),
        'girth': tree.get('girth', ''),
        'spread': tree.get('spread', ''),
        'route': survey_meta.get('route', ''),
        'diagnostician': survey_meta.get('diagnostician', ''),
        'date': format_date(survey_meta.get('date', '')),
    }

    for key, cell_addr in bi.items():
        value = mapping.get(key, '')
        if value != '':
            sheet[cell_addr] = value


def update_cell_checkbox(text: str, options: list, selected: str) -> str:
    """
    セル内文字列の中の □XXX を ■XXX に置換
    例: "□単独桝 □植栽帯 □緑地内 □その他" → "□単独桝 □植栽帯 ■緑地内 □その他"
    """
    if not selected or selected not in options:
        return text

    escaped = re.escape(selected)
    pattern = re.compile(r'□(\s*)' + escaped)
    if pattern.search(text):
        return pattern.sub(r'■\1' + selected, text)

    # フォールバック：単純置換
    return text.replace(f'□{selected}', f'■{selected}')


def write_cell_checkboxes(sheet: Worksheet, tree: dict, config: dict):
    """セル内チェックボックス（複数選択肢から1つを ■）"""
    cb_config = config['cell_checkboxes']

    mapping = {
        'planting_form': tree.get('plantingForm', ''),
        'stake': tree.get('stake', ''),
        'vitality_sei': tree.get('vitalitySei', ''),
        'vitality_kei': tree.get('vitalityKei', ''),
        'vitality_judgment': tree.get('vitalityJudgment', ''),
        'appearance_judgment': tree.get('appearanceJudgment', ''),
    }

    for key, def_obj in cb_config.items():
        raw_value = mapping.get(key, '')
        if not raw_value:
            continue

        # 入力値の変換
        convert = def_obj.get('convert_input')
        if convert == 'halfwidth_to_fullwidth':
            value = str(raw_value).translate(HALFWIDTH_TO_FULLWIDTH_DIGITS)
        elif convert == 'abc_to_fullwidth':
            value = ABC_TO_FULLWIDTH.get(raw_value, '')
        elif convert == 'abc_to_label':
            value = ABC_TO_VITALITY_LABEL.get(raw_value, '')
        else:
            value = raw_value

        if not value:
            continue

        cell_addr = def_obj['cell']
        original = sheet[cell_addr].value
        if original is None:
            continue

        new_text = update_cell_checkbox(str(original), def_obj['options'], value)
        if new_text != original:
            sheet[cell_addr] = new_text


def write_part_judgments(sheet: Worksheet, tree: dict, config: dict):
    """部位判定マトリクスの書き込み"""
    part_judgments = tree.get('partJudgments', {})
    if not part_judgments:
        return

    cells_map = config['part_judgment_cells']

    for part, judgment in part_judgments.items():
        if not judgment:
            continue
        if part not in cells_map:
            continue
        cell_addr = cells_map[part].get(judgment)
        if not cell_addr:
            continue
        original = sheet[cell_addr].value
        if original is None:
            continue
        # □ を ■ に置換
        new_text = str(original).replace('□', '■', 1)
        sheet[cell_addr] = new_text


def set_cell_checkbox(text: str, selected: str) -> str:
    """
    セル内の全チェックボックスを□にリセットし、指定された選択肢だけ■にする。
    例: "□なし□あり（" + selected="なし" → "■なし□あり（"
    例: "□なし□あり（" + selected="あり" → "□なし■あり（"
    """
    if not text:
        return text
    # 全て□にリセット
    result = text.replace('■', '□')
    # 指定された選択肢を■に
    escaped = re.escape(selected)
    pattern = re.compile(r'□(\s*)' + escaped)
    if pattern.search(result):
        result = pattern.sub(lambda m: '■' + m.group(1) + selected, result)
    return result


def get_negative_checkbox_text(checkbox_text: str) -> str:
    """正のチェックボックステキストから対応する否定テキストを返す"""
    if checkbox_text == '見えない':
        return '見える'
    return 'なし'


def write_diagnosis_checkboxes(sheet: Worksheet, tree: dict, config: dict):
    """部位診断のチェックボックス

    - デフォルトで全項目の「なし」に■をつける
    - 所見（メモ）に該当する病害チップがある場合、その部位×項目の「あり」に■をつけ「なし」を□にする
    """
    memo = tree.get('memo', '')
    parts_items = parse_memo_to_parts(memo) if memo else {'根元': [], '幹': [], '大枝': [], '_free': []}

    diagnosis_rows = config['diagnosis_rows']
    part_columns = config['part_columns']
    skip_items = set(config.get('skip_diagnosis_items', []))

    # メモから見つかった (部位, 正規化項目名) のセットを構築
    found_set = set()
    for part, items in parts_items.items():
        if part == '_free':
            continue
        for item_text in items:
            item_name = extract_diagnosis_item_name(item_text)
            if not item_name:
                continue
            normalized = item_name.replace('(', '（').replace(')', '）')
            if normalized in diagnosis_rows:
                row_def = diagnosis_rows[normalized]
                if isinstance(row_def, dict):
                    only_part = row_def.get('only_part')
                    # only_part指定がある場合はその部位に強制マッピング
                    if only_part:
                        found_set.add((only_part, normalized))
                    else:
                        found_set.add((part, normalized))
                else:
                    found_set.add((part, normalized))

    # 全診断行×部位を処理
    for item_name, row_def in diagnosis_rows.items():
        if item_name in skip_items:
            continue

        if isinstance(row_def, dict):
            row = row_def['row']
            only_part = row_def.get('only_part')
            checkbox_text = row_def.get('checkbox_text', 'あり')
            col_override = row_def.get('column_override')
        else:
            row = row_def
            only_part = None
            checkbox_text = 'あり'
            col_override = None

        negative_text = get_negative_checkbox_text(checkbox_text)

        # この項目が適用される部位リスト
        if only_part:
            parts_to_process = [only_part]
        else:
            parts_to_process = list(part_columns.keys())

        for part in parts_to_process:
            if only_part:
                col = col_override or part_columns.get(only_part, '')
            else:
                col = part_columns.get(part, '')

            if not col:
                continue

            cell_addr = f"{col}{row}"
            original = sheet[cell_addr].value
            if original is None:
                continue

            text = str(original)
            is_found = (part, item_name) in found_set

            if is_found:
                # 該当する病害が所見にある → あり（正）に■、なし（否）を□
                new_text = set_cell_checkbox(text, checkbox_text)
            else:
                # 該当する病害が所見にない → なし（否）に■、あり（正）を□
                new_text = set_cell_checkbox(text, negative_text)

            if new_text != str(original):
                sheet[cell_addr] = new_text


def write_shoken(sheet: Worksheet, tree: dict, config: dict):
    """所見欄に書き込み（部位別に整形）"""
    memo = tree.get('memo', '')
    if not memo:
        return

    parts = parse_memo_to_parts(memo)

    lines = []
    for part_label, key in [('根元', '根元'), ('幹', '幹'), ('枝', '大枝')]:
        items = parts.get(key, [])
        if items:
            lines.append(f"{part_label}：{('、').join(items)}")
    # 部位ラベルなしの自由記述
    if parts.get('_free'):
        lines.extend(parts['_free'])

    if not lines:
        return

    shoken_cfg = config['shoken']
    first_cell = shoken_cfg['first_cell']
    text = '\n'.join(lines)
    sheet[first_cell] = text

    # セル内改行のため wrap_text を有効化
    cell = sheet[first_cell]
    if cell.alignment:
        # 既存の alignment をコピーして wrap_text のみ更新
        from openpyxl.styles import Alignment
        new_alignment = Alignment(
            horizontal=cell.alignment.horizontal,
            vertical=cell.alignment.vertical or 'top',
            wrap_text=True,
            indent=cell.alignment.indent,
        )
        cell.alignment = new_alignment


def write_three_choice_circumference(sheet: Worksheet, tree: dict, config: dict):
    """3択項目（周囲長比率）の書き込み（v3.3）

    PWAの threeChoiceJudgments データに基づいて、該当セルの選択肢に■を付ける。
    """
    three_choice_config = config.get('three_choice_circumference')
    if not three_choice_config:
        return

    three_choice_data = tree.get('threeChoiceJudgments', {})

    # 内部キー → Excelラベルの対応
    KEY_TO_LABEL = {
        'none': 'なし',
        'less_third': '1/3未満',
        'more_third': '1/3以上',
    }

    for item_key in ['barkDeath', 'cavityShallow', 'cavityDeep']:
        if item_key not in three_choice_config:
            continue

        for part_key in ['root', 'trunk', 'branch']:
            cell_addr = three_choice_config[item_key].get(part_key)
            if not cell_addr:
                continue

            # データ取得（未入力なら 'none' をデフォルト）
            selected = three_choice_data.get(part_key, {}).get(item_key)
            if not selected:
                selected = 'none'

            # 有効なキーでなければ 'none' にフォールバック
            if selected not in KEY_TO_LABEL:
                selected = 'none'

            original = sheet[cell_addr].value
            if original is None:
                continue

            # セルテキスト内の選択された値だけ■、他は□
            new_text = set_cell_checkbox(str(original), KEY_TO_LABEL[selected])
            if new_text != str(original):
                sheet[cell_addr] = new_text


def write_overall_judgment(sheet: Worksheet, tree: dict, config: dict):
    """総合判定の書き込み（v3.2）"""
    # チェックボックス（G60: □Ａ：... □Ｂ１：... □Ｂ２：... □Ｃ：...）
    oj_config = config.get('overall_judgment')
    if oj_config:
        overall = tree.get('overallJudgment', '')
        if overall:
            value = ABC_TO_FULLWIDTH.get(overall, '')
        else:
            value = ''
        cell_addr = oj_config['cell']
        original = sheet[cell_addr].value
        if original is not None:
            new_text = update_cell_checkbox(str(original), oj_config['options'], value)
            if new_text != str(original):
                sheet[cell_addr] = new_text

    # 判定理由テキスト
    reason_cell = config.get('overall_reason_cell')
    if reason_cell:
        reason = tree.get('overallReason', '')
        if reason:
            sheet[reason_cell] = reason
            # wrap_text 有効化
            from openpyxl.styles import Alignment
            cell = sheet[reason_cell]
            if cell.alignment:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical or 'top',
                    wrap_text=True,
                    indent=cell.alignment.indent,
                )


# ===================================================================
# 写真の埋め込み
# ===================================================================

def cell_pixel_position(sheet: Worksheet, cell_addr: str) -> tuple:
    """セルの左上ピクセル座標を計算（おおよそ）"""
    cell = sheet[cell_addr]
    col_idx = cell.column  # 1-indexed
    row_idx = cell.row

    # 列幅をピクセルに変換して累積
    x = 0
    for c in range(1, col_idx):
        col_letter = get_column_letter(c)
        col_dim = sheet.column_dimensions.get(col_letter)
        width = col_dim.width if col_dim and col_dim.width else 8.43  # default
        # Excelの列幅 → pixel: width * 7 + 5 (おおよそ)
        x += width * 7 + 5

    # 行高をピクセルに変換して累積
    y = 0
    for r in range(1, row_idx):
        row_dim = sheet.row_dimensions.get(r)
        height = row_dim.height if row_dim and row_dim.height else 15  # default in pt
        y += pt_to_pixel(height)

    return x, y


def embed_photos(sheet: Worksheet, photos: list, config: dict):
    """写真をスロットに従って埋め込む"""
    if not photos:
        return

    slots = config['photo_slots']

    for photo in photos:
        label = photo.get('label')
        if not label or label not in slots:
            continue

        data_url = photo.get('dataUrl', '')
        m = re.match(r'^data:image/\w+;base64,(.+)$', data_url)
        if not m:
            continue

        try:
            img_bytes = base64.b64decode(m.group(1))

            # 全景写真(樹木全体)のアノテーション焼き込み
            annotations = photo.get('annotations', [])
            if label == '樹木全体' and annotations:
                annotated_buf = annotate_photo(io.BytesIO(img_bytes), annotations, output_format="PNG")
                pil_img = Image.open(annotated_buf)
            else:
                pil_img = Image.open(io.BytesIO(img_bytes))

            slot = slots[label]
            target_w = pt_to_pixel(slot['width_pt'])
            target_h = pt_to_pixel(slot['height_pt'])

            # アスペクト比を保つ場合
            if slot.get('keep_aspect_ratio'):
                pil_img.thumbnail((target_w, target_h), Image.Resampling.LANCZOS)
            else:
                pil_img = pil_img.resize((target_w, target_h), Image.Resampling.LANCZOS)

            # PIL → openpyxl Image
            buf = io.BytesIO()
            pil_img.save(buf, format='PNG')
            buf.seek(0)
            xl_img = XLImage(buf)
            xl_img.width = target_w
            xl_img.height = target_h

            # アンカーセルに配置
            anchor_cell = slot['anchor_cell']
            xl_img.anchor = anchor_cell
            sheet.add_image(xl_img)

        except Exception as e:
            warn(f"写真の埋め込みに失敗 (label={label}): {e}")


# ===================================================================
# メイン処理
# ===================================================================

def generate_karte(json_path: Path, output_path: Path, template_id: str):
    """カルテExcel生成のメイン処理"""

    # JSON読み込み
    log(f"JSONを読み込み中: {json_path}")
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    survey_meta = data.get('surveyMeta', {})
    trees = data.get('trees', [])

    if not trees:
        err("樹データが見つかりません")
        return

    log(f"樹木数: {len(trees)}")
    total_photos = sum(len(t.get('photos', [])) for t in trees)
    log(f"写真総数: {total_photos}")

    # テンプレート設定の読み込み
    log(f"テンプレート '{template_id}' を読み込み中...")
    config = load_template_config(template_id)
    template_path = get_template_excel_path(config)
    if not template_path.exists():
        err(f"テンプレートExcelが見つかりません: {template_path}")
        return

    # ワークブックを開く
    wb = load_workbook(template_path)
    template_sheet_name = config['sheet_name']

    if template_sheet_name not in wb.sheetnames:
        err(f"テンプレートシート '{template_sheet_name}' が見つかりません")
        return

    # 各樹について処理
    for i, tree in enumerate(trees):
        tree_no = tree.get('treeNumber') or str(i + 1)
        sheet_name = str(tree_no)[:31]  # Excelシート名は最大31文字

        # シート名重複対策
        original_name = sheet_name
        suffix = 2
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_name}_{suffix}"
            suffix += 1

        log(f"[{i+1}/{len(trees)}] 樹木 #{tree_no} ({tree.get('species', '')}) を処理中...")

        # シート複製
        new_sheet = copy_template_sheet(wb, template_sheet_name, sheet_name)

        # データ書き込み
        try:
            write_basic_info(new_sheet, tree, survey_meta, config)
            write_cell_checkboxes(new_sheet, tree, config)
            write_part_judgments(new_sheet, tree, config)
            write_diagnosis_checkboxes(new_sheet, tree, config)
            write_three_choice_circumference(new_sheet, tree, config)
            write_shoken(new_sheet, tree, config)
            write_overall_judgment(new_sheet, tree, config)

            # 写真埋め込み
            photos = tree.get('photos', [])
            if photos:
                embed_photos(new_sheet, photos, config)
        except Exception as e:
            err(f"樹木 #{tree_no} の処理中にエラー: {e}")
            import traceback
            traceback.print_exc()

    # 元のテンプレートシート（白紙）を削除
    if template_sheet_name in wb.sheetnames:
        del wb[template_sheet_name]

    # 保存
    log(f"カルテを保存中: {output_path}")
    wb.save(output_path)
    log("完了！")


def main():
    parser = argparse.ArgumentParser(
        description='街路樹現場調査JSONからカルテExcelを生成',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='例: python generate.py survey_2026-05-01.json'
    )
    parser.add_argument('json_file', help='PWAからエクスポートしたJSONファイル')
    parser.add_argument('--template', '-t', default='shibuya', help='テンプレートID (デフォルト: shibuya)')
    parser.add_argument('--output', '-o', help='出力ファイル名 (省略時はJSONと同名で.xlsx)')

    args = parser.parse_args()

    # 入力ファイルパス
    json_path = Path(args.json_file).resolve()
    if not json_path.exists():
        print(f"ERROR: JSONファイルが見つかりません: {json_path}", file=sys.stderr)
        sys.exit(1)

    # 出力ファイルパス
    if args.output:
        output_path = Path(args.output).resolve()
    else:
        # JSONと同じ場所に、同じベース名で .xlsx を作る
        base = json_path.stem
        # "survey_" プレフィックスは削除
        if base.startswith('survey_'):
            base = 'karte_' + base[7:]
        else:
            base = 'karte_' + base
        output_path = json_path.parent / f"{base}.xlsx"

    print()
    print("=" * 50)
    print("街路樹診断カルテ生成スクリプト")
    print("=" * 50)
    print()

    try:
        generate_karte(json_path, output_path, args.template)
    except Exception as e:
        print(f"\nERROR: 処理中にエラーが発生しました: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

    print()
    print("=" * 50)
    print(f"生成完了: {output_path}")
    print("=" * 50)
    print()


if __name__ == '__main__':
    main()
