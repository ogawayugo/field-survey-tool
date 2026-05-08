#!/usr/bin/env python3
"""
街路樹診断カルテ生成ツール（GUI版）

PWA「街路樹現場調査」からエクスポートしたJSONファイルから、
街路樹診断カルテのExcelファイルを生成するGUIアプリです。

使い方:
    1. 「カルテ生成.bat」をダブルクリック
    2. JSONファイルをウィンドウにドラッグ＆ドロップ（または[追加]ボタン）
    3. テンプレートを選択
    4. [カルテを生成]ボタンをクリック
    5. 同じフォルダに karte_xxx.xlsx が生成される
"""

import json
import os
import sys
import threading
import traceback
from datetime import datetime
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# generate.py の機能を再利用
# exe化されたかどうかで取得方法を分岐
if getattr(sys, 'frozen', False):
    # PyInstaller でexe化された場合：exeファイルのある場所
    SCRIPT_DIR = Path(sys.executable).parent
else:
    # 通常のPythonスクリプトとして実行された場合
    SCRIPT_DIR = Path(__file__).parent

sys.path.insert(0, str(SCRIPT_DIR))

try:
    from generate import (
        load_template_config,
        get_template_excel_path,
        copy_template_sheet,
        write_basic_info,
        write_cell_checkboxes,
        write_part_judgments,
        write_diagnosis_checkboxes,
        write_shoken,
        embed_photos,
    )
    from openpyxl import load_workbook
except ImportError as e:
    messagebox.showerror(
        "起動エラー",
        f"必要なモジュールが見つかりません。\n\n{e}\n\n"
        "generate.py がこのスクリプトと同じフォルダにあることを確認してください。"
    )
    sys.exit(1)


# ===================================================================
# テンプレートディレクトリ
# ===================================================================

TEMPLATES_DIR = SCRIPT_DIR / 'templates'


def discover_templates():
    """templates/ フォルダから利用可能なテンプレート一覧を取得"""
    templates = []
    if not TEMPLATES_DIR.exists():
        return templates

    for json_file in TEMPLATES_DIR.glob('*.json'):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
            template_id = json_file.stem
            templates.append({
                'id': template_id,
                'name': config.get('name', template_id),
            })
        except Exception:
            continue
    return templates


# ===================================================================
# カルテ生成のコア処理（GUI から呼ばれる）
# ===================================================================

def generate_karte_from_multiple_jsons(json_paths, output_path, template_id, progress_callback=None):
    """
    複数のJSONファイルから1つのカルテExcelを生成する。
    各JSON の trees を統合し、全樹を1ファイルに収める。

    progress_callback(current, total, message) で進捗を通知する。
    """
    # テンプレート設定の読み込み
    config = load_template_config(template_id)
    template_path = get_template_excel_path(config)
    if not template_path.exists():
        raise FileNotFoundError(f"テンプレートExcelが見つかりません: {template_path}")

    template_sheet_name = config['sheet_name']

    # 全JSONを読み込んで統合
    all_trees = []
    survey_meta = {}
    for json_path in json_paths:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        # 最初のJSONのsurveyMetaを使う（同じ調査の前提）
        if not survey_meta:
            survey_meta = data.get('surveyMeta', {})
        trees = data.get('trees', [])
        # JSON のファイル名から「担当者」を推測してメタ情報追加
        json_basename = Path(json_path).stem
        for tree in trees:
            tree['_source_file'] = json_basename
        all_trees.extend(trees)

    if not all_trees:
        raise ValueError("樹データが見つかりません")

    total = len(all_trees)
    if progress_callback:
        progress_callback(0, total, f"処理開始... 樹木{total}本")

    # ワークブックを開く
    wb = load_workbook(template_path)

    if template_sheet_name not in wb.sheetnames:
        raise ValueError(f"テンプレートシート '{template_sheet_name}' が見つかりません")

    # 各樹について処理
    for i, tree in enumerate(all_trees):
        tree_no = tree.get('treeNumber') or str(i + 1)
        species = tree.get('species', '')

        # シート名（重複対策）
        sheet_name = str(tree_no)[:31]
        original_name = sheet_name
        suffix = 2
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_name}_{suffix}"
            suffix += 1

        if progress_callback:
            progress_callback(i, total, f"処理中: 樹木 #{tree_no} ({species})")

        # シート複製
        new_sheet = copy_template_sheet(wb, template_sheet_name, sheet_name)

        # データ書き込み
        try:
            write_basic_info(new_sheet, tree, survey_meta, config)
            write_cell_checkboxes(new_sheet, tree, config)
            write_part_judgments(new_sheet, tree, config)
            write_diagnosis_checkboxes(new_sheet, tree, config)
            write_shoken(new_sheet, tree, config)

            photos = tree.get('photos', [])
            if photos:
                embed_photos(new_sheet, photos, config)
        except Exception as e:
            print(f"WARN: 樹木 #{tree_no} の処理中にエラー: {e}")
            traceback.print_exc()

    # 元のテンプレート（白紙）を削除
    if template_sheet_name in wb.sheetnames:
        del wb[template_sheet_name]

    if progress_callback:
        progress_callback(total, total, "Excelファイルを保存中...")

    wb.save(output_path)

    if progress_callback:
        progress_callback(total, total, "完了")


# ===================================================================
# GUI クラス
# ===================================================================

class KarteGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("街路樹診断カルテ生成ツール")
        self.root.geometry("680x520")
        self.root.minsize(580, 450)

        # 配色（紙のような暖色）
        self.bg_color = "#faf7f1"
        self.accent_color = "#1f3d2f"
        self.root.configure(bg=self.bg_color)

        self.json_files = []  # 選択されたJSONファイルのリスト
        self.templates = discover_templates()

        self._build_ui()
        self._setup_drag_drop()

    def _build_ui(self):
        # メインフレーム
        main = tk.Frame(self.root, bg=self.bg_color, padx=20, pady=20)
        main.pack(fill='both', expand=True)

        # タイトル
        title = tk.Label(
            main, text="街路樹診断カルテ生成ツール",
            font=("Yu Gothic", 14, "bold"),
            bg=self.bg_color, fg=self.accent_color
        )
        title.pack(anchor='w', pady=(0, 4))

        subtitle = tk.Label(
            main, text="PWAでエクスポートしたJSONファイルから、診断カルテExcelを生成します",
            font=("Yu Gothic", 9),
            bg=self.bg_color, fg="#6b5a3e"
        )
        subtitle.pack(anchor='w', pady=(0, 20))

        # ─── JSONファイルセクション ───
        json_label = tk.Label(
            main, text="① JSONファイル（複数可、ドラッグ＆ドロップ可）",
            font=("Yu Gothic", 10, "bold"),
            bg=self.bg_color, fg="#3a3a3a"
        )
        json_label.pack(anchor='w')

        # ファイルリスト
        list_frame = tk.Frame(main, bg=self.bg_color)
        list_frame.pack(fill='x', pady=(6, 0))

        self.file_listbox = tk.Listbox(
            list_frame, height=6,
            font=("Yu Gothic", 10),
            bg="white", fg="#3a3a3a",
            selectmode='extended',
            relief='solid', borderwidth=1
        )
        self.file_listbox.pack(side='left', fill='both', expand=True)

        scrollbar = tk.Scrollbar(list_frame, command=self.file_listbox.yview)
        scrollbar.pack(side='right', fill='y')
        self.file_listbox.config(yscrollcommand=scrollbar.set)

        # ボタン群
        btn_frame = tk.Frame(main, bg=self.bg_color)
        btn_frame.pack(fill='x', pady=(6, 16))

        self._make_button(btn_frame, "ファイルを追加", self._add_files).pack(side='left', padx=(0, 6))
        self._make_button(btn_frame, "選択を削除", self._remove_selected).pack(side='left', padx=(0, 6))
        self._make_button(btn_frame, "全てクリア", self._clear_all).pack(side='left')

        # ─── テンプレートセクション ───
        tpl_label = tk.Label(
            main, text="② テンプレート",
            font=("Yu Gothic", 10, "bold"),
            bg=self.bg_color, fg="#3a3a3a"
        )
        tpl_label.pack(anchor='w', pady=(8, 6))

        if self.templates:
            template_names = [t['name'] for t in self.templates]
            self.template_var = tk.StringVar(value=template_names[0])
            self.template_combo = ttk.Combobox(
                main, textvariable=self.template_var,
                values=template_names, state='readonly',
                font=("Yu Gothic", 10),
            )
            self.template_combo.pack(fill='x')
        else:
            self.template_var = tk.StringVar(value="（テンプレートが見つかりません）")
            tk.Label(main, textvariable=self.template_var,
                     bg="white", relief='solid', borderwidth=1,
                     font=("Yu Gothic", 10), fg="red", anchor='w', padx=8, pady=4
                     ).pack(fill='x')

        # ─── 出力ファイル名 ───
        out_label = tk.Label(
            main, text="③ 出力ファイル名",
            font=("Yu Gothic", 10, "bold"),
            bg=self.bg_color, fg="#3a3a3a"
        )
        out_label.pack(anchor='w', pady=(16, 6))

        out_frame = tk.Frame(main, bg=self.bg_color)
        out_frame.pack(fill='x')

        today = datetime.now().strftime("%Y-%m-%d")
        self.output_var = tk.StringVar(value=f"karte_{today}.xlsx")
        self.output_entry = tk.Entry(
            out_frame, textvariable=self.output_var,
            font=("Yu Gothic", 10),
            relief='solid', borderwidth=1
        )
        self.output_entry.pack(side='left', fill='x', expand=True, padx=(0, 6))

        self._make_button(out_frame, "保存先を指定", self._choose_output).pack(side='right')

        # ─── 生成ボタン ───
        generate_btn = tk.Button(
            main, text="カルテを生成",
            command=self._on_generate,
            font=("Yu Gothic", 11, "bold"),
            bg=self.accent_color, fg="white",
            activebackground="#0f2d20", activeforeground="white",
            relief='flat', cursor='hand2',
            padx=20, pady=10
        )
        generate_btn.pack(fill='x', pady=(20, 6))
        self.generate_btn = generate_btn

        # ─── 進捗 ───
        self.progress_var = tk.StringVar(value="")
        self.progress_label = tk.Label(
            main, textvariable=self.progress_var,
            font=("Yu Gothic", 9),
            bg=self.bg_color, fg="#6b5a3e", anchor='w'
        )
        self.progress_label.pack(fill='x', pady=(4, 0))

        self.progress_bar = ttk.Progressbar(main, mode='determinate', length=100)
        self.progress_bar.pack(fill='x', pady=(2, 0))

    def _make_button(self, parent, text, command):
        return tk.Button(
            parent, text=text, command=command,
            font=("Yu Gothic", 9),
            bg="white", fg="#3a3a3a",
            activebackground="#f0ede5",
            relief='solid', borderwidth=1, cursor='hand2',
            padx=10, pady=4
        )

    def _setup_drag_drop(self):
        """ドラッグ＆ドロップを有効化（tkinterDnD2が利用可能なら）"""
        try:
            from tkinterdnd2 import DND_FILES, TkinterDnD
            # ドラッグ&ドロップは optional
            self.file_listbox.drop_target_register(DND_FILES)
            self.file_listbox.dnd_bind('<<Drop>>', self._on_drop)
        except ImportError:
            # tkinterDnD2 がインストールされていない場合は何もしない
            # （ファイル選択ボタンで代用）
            pass

    def _on_drop(self, event):
        """ドラッグ＆ドロップでファイル追加"""
        # event.data はパスのスペース区切り（クォート付きの場合あり）
        files = self.root.tk.splitlist(event.data)
        added = 0
        for f in files:
            f = f.strip('{}')  # 中括弧クォート除去
            if f.lower().endswith('.json') and f not in self.json_files:
                self.json_files.append(f)
                self.file_listbox.insert('end', os.path.basename(f))
                added += 1

    def _add_files(self):
        """ファイル選択ダイアログ"""
        files = filedialog.askopenfilenames(
            title="JSONファイルを選択",
            filetypes=[("JSON ファイル", "*.json"), ("すべてのファイル", "*.*")]
        )
        for f in files:
            if f not in self.json_files:
                self.json_files.append(f)
                self.file_listbox.insert('end', os.path.basename(f))

    def _remove_selected(self):
        """選択したファイルを削除"""
        selected = list(self.file_listbox.curselection())
        for i in reversed(selected):
            self.file_listbox.delete(i)
            del self.json_files[i]

    def _clear_all(self):
        """全クリア"""
        self.file_listbox.delete(0, 'end')
        self.json_files.clear()

    def _choose_output(self):
        """出力先を指定"""
        # JSONファイルがあれば、その親フォルダをデフォルトに
        initial_dir = None
        if self.json_files:
            initial_dir = os.path.dirname(self.json_files[0])

        f = filedialog.asksaveasfilename(
            title="保存先を選択",
            defaultextension=".xlsx",
            initialdir=initial_dir,
            initialfile=self.output_var.get(),
            filetypes=[("Excel ファイル", "*.xlsx")]
        )
        if f:
            self.output_var.set(f)

    def _on_generate(self):
        """生成ボタン押下"""
        if not self.json_files:
            messagebox.showwarning("ファイル未選択", "JSONファイルを追加してください。")
            return

        if not self.templates:
            messagebox.showerror("テンプレートエラー", "テンプレートが見つかりません。")
            return

        # テンプレートID取得
        selected_name = self.template_var.get()
        template_id = None
        for t in self.templates:
            if t['name'] == selected_name:
                template_id = t['id']
                break
        if not template_id:
            messagebox.showerror("エラー", "テンプレートを選択してください。")
            return

        # 出力パス取得
        output_name = self.output_var.get().strip()
        if not output_name:
            messagebox.showerror("エラー", "出力ファイル名を入力してください。")
            return

        # 出力先パスを決定
        if os.path.isabs(output_name):
            output_path = Path(output_name)
        else:
            # JSONファイルと同じフォルダに出力
            first_json_dir = Path(self.json_files[0]).parent
            output_path = first_json_dir / output_name

        # 別スレッドで実行（UI 凍結防止）
        self.generate_btn.config(state='disabled')
        thread = threading.Thread(
            target=self._run_generate,
            args=(self.json_files.copy(), output_path, template_id),
            daemon=True
        )
        thread.start()

    def _run_generate(self, json_paths, output_path, template_id):
        """別スレッドで実行される生成処理"""
        try:
            generate_karte_from_multiple_jsons(
                json_paths, output_path, template_id,
                progress_callback=self._update_progress
            )
            self.root.after(0, lambda: self._on_complete(output_path))
        except Exception as e:
            error_msg = f"{e}\n\n{traceback.format_exc()}"
            self.root.after(0, lambda: self._on_error(error_msg))

    def _update_progress(self, current, total, message):
        """進捗更新（メインスレッドで実行）"""
        def update():
            if total > 0:
                percent = int(current / total * 100)
                self.progress_bar['value'] = percent
            self.progress_var.set(f"{message} ({current}/{total})")

        self.root.after(0, update)

    def _on_complete(self, output_path):
        """完了"""
        self.generate_btn.config(state='normal')
        self.progress_var.set(f"完了: {output_path}")

        # 成功ダイアログ
        result = messagebox.askyesno(
            "生成完了",
            f"カルテファイルを生成しました：\n{output_path}\n\n保存先のフォルダを開きますか？"
        )
        if result:
            # 保存先フォルダを開く
            try:
                os.startfile(str(output_path.parent))
            except AttributeError:
                # macOS / Linux 対応
                import subprocess
                subprocess.run(['open' if sys.platform == 'darwin' else 'xdg-open', str(output_path.parent)])

    def _on_error(self, error_msg):
        """エラー"""
        self.generate_btn.config(state='normal')
        self.progress_var.set("エラー")
        self.progress_bar['value'] = 0
        messagebox.showerror("生成エラー", f"カルテ生成中にエラーが発生しました：\n\n{error_msg}")


def main():
    # tkinterdnd2 が使えればドラッグ&ドロップ対応のルートを使う
    try:
        from tkinterdnd2 import TkinterDnD
        root = TkinterDnD.Tk()
    except ImportError:
        root = tk.Tk()

    app = KarteGeneratorApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
